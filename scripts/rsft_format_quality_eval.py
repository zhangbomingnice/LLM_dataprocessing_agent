#!/usr/bin/env python3
"""
RSFT 长回答「格式质量」批量评测 → Excel

支持：
  - 单个 JSONL：每行 id + instruction + prediction，作为唯一候选 C1
  - 最多 8 个 JSONL：按 id 对齐合并，文件名顺序对应 C1…C8

输入 JSONL 字段（每行）：
  - id（或 prompt_id）
  - instruction（题目）
  - prediction（回答文本；可与 response 二选一）
  - subset（可选）

用法示例：
  # 仅评测当前单文件（每行 1 个回答 → 报告中只有 C1 列）
  python scripts/rsft_format_quality_eval.py \\
    --inputs "D:/llm_posttrain/outputs/rsft_smoke_v1_eval_public_predictions.jsonl" \\
    --output "D:/llm_posttrain/outputs/format_scores.xlsx"

  # 八个模型输出（8 个文件，相同 id 对齐）
  python scripts/rsft_format_quality_eval.py \\
    --inputs m1.jsonl m2.jsonl ... m8.jsonl \\
    --output scores.xlsx

环境变量：MINIMAX_API_KEY 或 OPENAI_API_KEY；可用 --api-key 覆盖。
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from pathlib import Path

# 项目根目录
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv
from openai import AsyncOpenAI
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

load_dotenv(ROOT / ".env")
console = Console()
logger = logging.getLogger(__name__)

PROMPT_PATH = ROOT / "configs" / "judge_prompts" / "rsft_format_quality_zh.txt"


def _truncate(s: str, max_chars: int) -> str:
    s = s or ""
    if len(s) <= max_chars:
        return s
    head = max_chars * 65 // 100
    tail = max_chars - head - 80
    if tail < 0:
        return s[:max_chars]
    return s[:head] + "\n\n[…中间省略…]\n\n" + s[-tail:]


def _load_jsonl(path: Path) -> list[dict]:
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def _row_key(r: dict) -> str:
    if "id" in r:
        return str(r["id"])
    if "prompt_id" in r:
        return str(r["prompt_id"])
    raise KeyError("行中缺少 id 或 prompt_id")


def _row_instruction(r: dict) -> str:
    return (r.get("instruction") or r.get("text") or r.get("question") or "").strip()


def _row_response(r: dict) -> str:
    return (r.get("prediction") or r.get("response") or r.get("answer") or "").strip()


def _predictions_from_row(r: dict) -> list[str]:
    """单行多个回答：predictions / candidates 为字符串列表；否则用单个 prediction。"""
    raw = r.get("predictions") or r.get("candidates")
    if isinstance(raw, list):
        return [str(x or "").strip() for x in raw]
    return [_row_response(r)]


def merge_by_id(paths: list[Path]) -> list[dict]:
    """
    返回 [{ "id", "subset", "instruction", "predictions": [p1,...,pK] }, ...]

    - 仅 1 个文件：每行可用 predictions 列表（8 个回答），否则用 prediction 单字段（K=1）。
    - 多个文件：按 id 取交集，第 i 个文件提供第 i 个回答，K=文件数。
    """
    if len(paths) == 1:
        merged = []
        for r in _load_jsonl(paths[0]):
            preds = _predictions_from_row(r)
            if len(preds) == 1 and not preds[0]:
                preds = [_row_response(r)]
            merged.append({
                "id": _row_key(r),
                "subset": r.get("subset", ""),
                "instruction": _row_instruction(r),
                "predictions": preds,
            })
        return merged

    per_file: list[dict[str, dict]] = []
    for p in paths:
        m: dict[str, dict] = {}
        for r in _load_jsonl(p):
            k = _row_key(r)
            m[k] = r
        per_file.append(m)

    common_ids = set(per_file[0].keys())
    for m in per_file[1:]:
        common_ids &= set(m.keys())

    missing = set(per_file[0].keys()) - common_ids
    if missing and len(paths) > 1:
        logger.warning("有 %d 个 id 未在所有输入文件中出现，已跳过", len(missing))

    merged = []

    def _sort_id(x: str) -> tuple:
        s = str(x)
        return (0, int(s)) if s.isdigit() else (1, s)

    for pid in sorted(common_ids, key=_sort_id):
        instrs = [_row_instruction(per_file[i][pid]) for i in range(len(paths))]
        base = instrs[0]
        if len(set(instrs)) > 1:
            logger.warning("id=%s 在不同文件中 instruction 不一致，使用第一个文件中的文本", pid)
        preds = [_row_response(per_file[i][pid]) for i in range(len(paths))]
        subset = per_file[0][pid].get("subset", "")
        merged.append({
            "id": pid,
            "subset": subset,
            "instruction": base,
            "predictions": preds,
        })
    return merged


def max_candidate_count(merged: list[dict]) -> int:
    return max(len(m["predictions"]) for m in merged) if merged else 0


def build_user_message(item: dict, max_chars_per_cand: int) -> str:
    lines = ["## 题目（Instruction）", item["instruction"], "", "## 候选回答"]
    for i, pred in enumerate(item["predictions"], start=1):
        label = f"C{i}"
        body = _truncate(pred, max_chars_per_cand)
        lines.append(f"### {label}")
        lines.append(body if body else "（空回答）")
        lines.append("")
    return "\n".join(lines)


def _parse_json_loose(text: str) -> dict:
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    m = re.search(r"\{[\s\S]*\}\s*$", text)
    if m:
        return json.loads(m.group())
    m2 = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if m2:
        return json.loads(m2.group(1).strip())
    raise ValueError("无法解析 JSON")


class FormatEvalRunner:
    def __init__(
        self,
        api_key: str,
        base_url: str,
        model: str,
        system_prompt: str,
        max_tokens: int = 8192,
    ):
        self._client = AsyncOpenAI(api_key=api_key, base_url=base_url)
        self.model = model
        self.system_prompt = system_prompt
        self.max_tokens = max_tokens

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=40),
        retry=retry_if_exception_type(Exception),
        reraise=True,
    )
    async def _one(self, user_msg: str, temperature: float = 0.2) -> dict:
        resp = await self._client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "system", "content": self.system_prompt},
                {"role": "user", "content": user_msg},
            ],
            temperature=temperature,
            max_tokens=self.max_tokens,
        )
        raw = (resp.choices[0].message.content or "").strip()
        return _parse_json_loose(raw)

def write_excel(path: Path, rows: list[dict], n_candidates: int) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "format_scores"

    base_headers = [
        "id", "subset", "instruction_preview",
        "best", "ranking", "verdict_summary", "error",
    ]
    sub_dims = [
        ("response_mode", "回答模式20"),
        ("structural_completeness", "结构完整25"),
        ("organization", "组织条理20"),
        ("fluency", "流畅度15"),
        ("non_repetition", "去重复15"),
        ("task_fit", "题型适配5"),
        ("total", "总分100"),
    ]

    headers = list(base_headers)
    for i in range(1, n_candidates + 1):
        prefix = f"C{i}"
        for key, zh in sub_dims:
            headers.append(f"{prefix}_{zh}")

    ws.append(headers)

    for row in rows:
        base = [
            row.get("id"),
            row.get("subset"),
            (row.get("instruction_preview") or "")[:500],
            row.get("best"),
            row.get("ranking"),
            row.get("verdict_summary"),
            row.get("error"),
        ]
        scores_flat = row.get("scores_flat") or {}
        rest = [scores_flat.get(h, "") for h in headers[len(base) :]]
        ws.append(base + rest)

    # 简单列宽
    for col_idx, h in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(45, max(10, len(str(h)) + 2))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    console.print(f"[green]已写入[/green] {path.resolve()}")


def flatten_scores(parsed: dict | None, n_cand: int) -> dict[str, object]:
    out: dict[str, object] = {}
    if not parsed:
        return out
    evs = parsed.get("evaluations") or []
    by_label = {e.get("label"): e for e in evs if isinstance(e, dict)}
    for i in range(1, n_cand + 1):
        label = f"C{i}"
        e = by_label.get(label) or (evs[i - 1] if i - 1 < len(evs) else None)
        if not isinstance(e, dict):
            continue
        out[f"C{i}_回答模式20"] = e.get("response_mode", "")
        out[f"C{i}_结构完整25"] = e.get("structural_completeness", "")
        out[f"C{i}_组织条理20"] = e.get("organization", "")
        out[f"C{i}_流畅度15"] = e.get("fluency", "")
        out[f"C{i}_去重复15"] = e.get("non_repetition", "")
        out[f"C{i}_题型适配5"] = e.get("task_fit", "")
        out[f"C{i}_总分100"] = e.get("total", "")
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="RSFT 格式质量评测 → Excel")
    parser.add_argument(
        "--inputs", "-i", nargs="+", required=True,
        help="一个或多个 JSONL，按顺序对应 C1、C2…（最多 8 个）；多文件时按 id 取交集对齐",
    )
    parser.add_argument(
        "--output", "-o", required=True,
        help="输出 .xlsx 路径",
    )
    parser.add_argument("--api-key", default="", help="覆盖环境变量")
    parser.add_argument(
        "--base-url",
        default=os.getenv("MINIMAX_API_BASE")
        or os.getenv("OPENAI_BASE_URL")
        or "https://api.minimax.io/v1",
    )
    parser.add_argument(
        "--model",
        default=os.getenv("MINIMAX_MODEL")
        or os.getenv("MODEL_NAME")
        or "MiniMax-M2.7",
    )
    parser.add_argument("--concurrency", type=int, default=4)
    parser.add_argument("--max-tokens", type=int, default=8192)
    parser.add_argument(
        "--max-chars-per-cand",
        type=int,
        default=2800,
        help="每个候选回答写入 prompt 的最大字符数（8 候选时建议 2000~3500）",
    )
    parser.add_argument(
        "--prompt-file",
        type=str,
        default=str(PROMPT_PATH),
        help="Judge 系统提示词路径",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.WARNING)

    paths = [Path(p).resolve() for p in args.inputs]
    if len(paths) > 8:
        console.print("[red]最多 8 个输入文件[/red]")
        sys.exit(1)
    for p in paths:
        if not p.exists():
            console.print(f"[red]文件不存在: {p}[/red]")
            sys.exit(1)

    api_key = args.api_key or os.getenv("MINIMAX_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        console.print("[red]请设置 MINIMAX_API_KEY 或 OPENAI_API_KEY，或使用 --api-key[/red]")
        sys.exit(1)

    prompt_path = Path(args.prompt_file)
    system_prompt = prompt_path.read_text(encoding="utf-8")

    merged = merge_by_id(paths)
    if not merged:
        console.print("[red]没有可对齐的数据行[/red]")
        sys.exit(1)

    n_cand = max_candidate_count(merged)
    if n_cand > 8:
        console.print("[red]每题候选回答超过 8 个（当前 {}），请减少 predictions 数量或合并文件数[/red]".format(n_cand))
        sys.exit(1)
    if n_cand < 1:
        console.print("[red]未找到任何回答文本[/red]")
        sys.exit(1)

    console.print(
        f"加载 [cyan]{len(merged)}[/cyan] 条题目；"
        f"表格将包含 [cyan]{n_cand}[/cyan] 组分数列 (C1…C{n_cand})",
    )

    runner = FormatEvalRunner(
        api_key=api_key,
        base_url=args.base_url.rstrip("/"),
        model=args.model,
        system_prompt=system_prompt,
        max_tokens=args.max_tokens,
    )

    async def run_all():
        out_rows = []
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("评测中", total=len(merged))
            sem = asyncio.Semaphore(args.concurrency)

            async def one(it: dict):
                async with sem:
                    try:
                        msg = build_user_message(it, args.max_chars_per_cand)
                        data = await runner._one(msg)
                        err = None
                    except Exception as e:
                        data = None
                        err = str(e)
                    progress.advance(task)
                    return it, data, err

            raw_results = await asyncio.gather(*[one(it) for it in merged])

        for it, data, err in raw_results:
            row = {
                "id": it["id"],
                "subset": it.get("subset", ""),
                "instruction_preview": it["instruction"][:200],
                "best": (data or {}).get("best", "") if data else "",
                "ranking": json.dumps((data or {}).get("ranking", []), ensure_ascii=False) if data else "",
                "verdict_summary": (data or {}).get("verdict_summary", "") if data else "",
                "error": err or "",
                "scores_flat": flatten_scores(data, n_cand),
            }
            out_rows.append(row)
        return out_rows

    rows = asyncio.run(run_all())
    write_excel(Path(args.output), rows, n_cand)
    ok = sum(1 for r in rows if not r.get("error"))
    console.print(f"完成: [green]{ok}[/green] / {len(rows)} 条成功")


if __name__ == "__main__":
    main()
